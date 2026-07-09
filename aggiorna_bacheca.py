#!/usr/bin/env python3
"""Costruisce i dati della pagina "Bacheca" del sito dai tre Excel della lega:

  - "ALBO D'ORO & PREMI INDIVIDUALI - FantaNBA Hezonja.xlsx"  (albo d'oro + premi)
  - "Il file dei Record del FantaNBA.xlsx"                    (record)
  - "FantaNBA.xlsx"  foglio "Bacheca"                        (maglie ritirate)

Scrive il blocco `window.BACHECA = {...};` dentro index.html (e in bacheca.js).

Uso:  python aggiorna_bacheca.py
Requisiti:  pip install openpyxl
"""
import json
import re
import sys

from openpyxl import load_workbook

ALBO = "ALBO D'ORO & PREMI INDIVIDUALI - FantaNBA Hezonja.xlsx"
RECORD = "Il file dei Record del FantaNBA.xlsx"
LEGA = "FantaNBA.xlsx"

SEASON_LONG = re.compile(r"^\d{4}\s*-\s*\d{4}$")   # 2018-2019
SEASON_SHORT = re.compile(r"^'?\d{2}\s*-\s*'?\d{2}$")   # 18-19 / '18-'19


def clean(v):
    """Cella -> stringa a riga singola, spazi normalizzati ('' se vuota)."""
    if v is None:
        return ""
    return " ".join(str(v).split())


def multiline(v):
    """Cella -> lista di righe non vuote."""
    if v is None:
        return []
    return [r.strip() for r in str(v).splitlines() if r.strip()]


def sheet_rows(wb, name):
    return list(wb[name].iter_rows(values_only=True))


def cell(row, i):
    return row[i] if i < len(row) else None


# ---------------- Albo d'oro: Finals ----------------

def estrai_finals(wb):
    out = []
    for r in sheet_rows(wb, "ALBO DORO FINALS"):
        stag = clean(cell(r, 0))
        if not SEASON_LONG.match(stag):
            continue
        champ = clean(cell(r, 1))
        if champ in ("", "-"):
            champ = ""
        mvp = re.sub(r"^MVP:\s*", "", clean(cell(r, 6)))
        out.append({
            "season": stag,
            "champion": champ,
            "series": clean(cell(r, 3)),
            "mvp": mvp,
        })
    return out


# ---------------- Albo d'oro: Conference & Division ----------------

def estrai_conf_div(wb):
    conf, div = [], []
    for r in sheet_rows(wb, "ALBO DORO CONFERENCE - DIVISION"):
        s_conf = clean(cell(r, 0))
        if SEASON_SHORT.match(s_conf):
            conf.append({
                "season": s_conf.replace("'", ""),
                "nord": clean(cell(r, 1)),
                "sud": clean(cell(r, 2)),
            })
        s_div = clean(cell(r, 4))
        if SEASON_SHORT.match(s_div):
            div.append({
                "season": s_div.replace("'", ""),
                "nordEst": clean(cell(r, 5)),
                "nordWest": clean(cell(r, 7)),
                "centroSud": clean(cell(r, 9)),
                "sudSud": clean(cell(r, 11)),
            })
    return conf, div


# ---------------- Premi individuali ----------------

def estrai_premi(wb):
    premi, cur = [], None
    LABELS = {"MVP:": "mvp", "MIP:": "mip", "ROY:": "roy", "COY:": "coy"}
    for r in sheet_rows(wb, "PREMI INDIVIDUALI"):
        s = clean(cell(r, 0))
        label = clean(cell(r, 1))
        if label not in LABELS:
            continue
        if SEASON_SHORT.match(s):
            cur = {"season": s.replace("'", ""), "mvp": "", "mip": "", "roy": "", "coy": ""}
            premi.append(cur)
        if cur is None:
            continue
        vincitore = clean(cell(r, 2))
        team = re.sub(r"^-\s*", "", clean(cell(r, 6)))
        if vincitore.lower() == "not assigned":
            cur[LABELS[label]] = ""
        else:
            cur[LABELS[label]] = f"{vincitore} — {team}" if team else vincitore
    return premi


# ---------------- Record ----------------

def estrai_record():
    wb = load_workbook(RECORD, read_only=True, data_only=True)
    squadra, individuali = [], []
    for r in sheet_rows(wb, "Foglio1"):
        lab_s = clean(cell(r, 1))
        val_s = clean(cell(r, 3))
        if lab_s.endswith(":") and val_s:
            squadra.append({
                "label": lab_s.rstrip(":"),
                "value": val_s,
                "team": clean(cell(r, 4)),
                "season": clean(cell(r, 5)),
            })
        lab_i = clean(cell(r, 8))
        val_i = clean(cell(r, 10))
        if lab_i.endswith(":") and val_i:
            individuali.append({
                "label": lab_i.rstrip(":"),
                "value": val_i,
                "match": clean(cell(r, 11)),
                "player": clean(cell(r, 12)),
                "season": clean(cell(r, 13)),
            })
    wb.close()
    return squadra, individuali


# ---------------- Bacheca: maglie ritirate ----------------

def estrai_maglie():
    wb = load_workbook(LEGA, read_only=True, data_only=True)
    out = []
    for r in sheet_rows(wb, "Bacheca"):
        righe = multiline(cell(r, 8))
        if not righe or righe[0].lower().startswith("maglie"):
            continue
        player = righe[0]
        season = next((x.split(":", 1)[1].strip() for x in righe if x.lower().startswith("season")), "")
        note = next((x for x in righe[1:] if not x.lower().startswith("season")), "")
        out.append({"player": player, "season": season, "note": note})
    wb.close()
    return out


def medagliere_nba(finals):
    conta = {}
    for f in finals:
        if f["champion"]:
            conta[f["champion"]] = conta.get(f["champion"], 0) + 1
    return [{"team": t, "titoli": n}
            for t, n in sorted(conta.items(), key=lambda kv: (-kv[1], kv[0]))]


def main():
    wb = load_workbook(ALBO, read_only=True, data_only=True)
    finals = estrai_finals(wb)
    conf, div = estrai_conf_div(wb)
    premi = estrai_premi(wb)
    wb.close()
    rec_sq, rec_ind = estrai_record()

    bacheca = {
        "finals": finals,
        "conference": conf,
        "division": div,
        "premi": premi,
        "recordSquadra": rec_sq,
        "recordIndividuali": rec_ind,
        "maglie": estrai_maglie(),
        "medagliere": medagliere_nba(finals),
    }

    payload = "window.BACHECA = " + json.dumps(bacheca, ensure_ascii=False) + ";"
    with open("bacheca.js", "w", encoding="utf-8") as f:
        f.write(payload)

    try:
        html = open("index.html", encoding="utf-8").read()
        html, n = re.subn(r"window\.BACHECA = \{.*?\};", payload, html, count=1, flags=re.S)
        if n:
            open("index.html", "w", encoding="utf-8").write(html)
            print("index.html: blocco BACHECA aggiornato.")
        else:
            print("! In index.html non ho trovato 'window.BACHECA = {...};' da sostituire.")
    except FileNotFoundError:
        pass

    print(f"OK: {len(finals)} finals, {len(conf)} conference, {len(div)} division, "
          f"{len(premi)} stagioni premi, {len(rec_sq)}+{len(rec_ind)} record, "
          f"{len(bacheca['maglie'])} maglie ritirate.")


if __name__ == "__main__":
    main()
