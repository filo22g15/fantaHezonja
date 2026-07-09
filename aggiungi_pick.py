#!/usr/bin/env python3
"""Importa (una tantum) le scelte al draft dai fogli squadra dell'Excel.

ATTENZIONE: da quando le pick si gestiscono dal pannello Admin del sito, questo
è solo un seeder iniziale. Rilanciarlo SOVRASCRIVE con l'Excel le pick già
presenti nel sito (inclusi gli scambi fatti dall'Admin). Se ci sono già pick,
lo script chiede conferma; con --forza salta la domanda.

Legge i dati CORRENTI da index.html (che contiene già stipendi e rinnovi 'pnd'),
aggiunge solo le pick e riscrive data.js + index.html.

Uso:  python aggiungi_pick.py [FantaNBA.xlsx] [--forza]
Requisiti:  pip install openpyxl
"""
import json
import re
import sys

from openpyxl import load_workbook

FORZA = "--forza" in sys.argv
args = [a for a in sys.argv[1:] if not a.startswith("--")]
XLSX = args[0] if args else "FantaNBA.xlsx"
ANNO_MIN = 2026   # mostra solo le pick da quest'anno in poi (le precedenti sono storiche)

# "scelta 1° giro BANDITS 2026"  ->  round=1, origine=BANDITS, anno=2026
RE_PICK = re.compile(r"scelta\s*(\d+)\s*°?\s*giro\s+(.+?)\s+((?:19|20)\d{2})", re.I)


def estrai_pick(ws):
    """Ritorna la lista di pick future (dedup, ordinate) trovate nel foglio."""
    trovate, viste = [], set()
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if not isinstance(v, str) or "scelta" not in v.lower():
                continue
            testo = " ".join(v.replace("�", "°").split())   # ripulisce il mojibake del '°'
            m = RE_PICK.search(testo)
            if not m:
                continue
            rd, origine, anno = int(m.group(1)), m.group(2).strip(), int(m.group(3))
            if anno < ANNO_MIN:
                continue
            chiave = (rd, origine.upper(), anno)
            if chiave in viste:
                continue
            viste.add(chiave)
            trovate.append({"y": anno, "rd": rd, "from": origine})
    trovate.sort(key=lambda p: (p["y"], p["rd"], p["from"]))
    return trovate


def main():
    wb = load_workbook(XLSX, read_only=True, data_only=True)

    html = open("index.html", encoding="utf-8").read()
    m = re.search(r"window\.LEAGUE = (\{.*?\});", html, re.S)
    if not m:
        sys.exit("Blocco window.LEAGUE non trovato in index.html")
    data = json.loads(m.group(1))

    gia_presenti = sum(len(t.get("picks") or []) for t in data["teams"])
    if gia_presenti and not FORZA:
        print(f"Nel sito ci sono già {gia_presenti} scelte (gestite dall'Admin).")
        risposta = input("Sovrascriverle con quelle dell'Excel? Scrivi SI per procedere: ")
        if risposta.strip().upper() != "SI":
            sys.exit("Annullato: nessuna modifica.")

    tot = 0
    for t in data["teams"]:
        sn = t.get("sheet")
        if not sn or sn not in wb.sheetnames:
            t["picks"] = []
            print(f"  ! foglio non trovato per {t['name']} (sheet={sn!r})")
            continue
        t["picks"] = estrai_pick(wb[sn])
        tot += len(t["picks"])
        print(f"  {t['name']}: {len(t['picks'])} scelte")

    payload = "window.LEAGUE = " + json.dumps(data, ensure_ascii=False) + ";"
    open("data.js", "w", encoding="utf-8").write(payload)
    html2 = re.sub(r"window\.LEAGUE = \{.*?\};", lambda _: payload, html, count=1, flags=re.S)
    open("index.html", "w", encoding="utf-8").write(html2)
    print(f"\nOK: {tot} scelte totali aggiunte. Ricarica index.html (o data.js) su GitHub.")


if __name__ == "__main__":
    main()
